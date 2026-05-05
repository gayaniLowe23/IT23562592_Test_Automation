from playwright.sync_api import sync_playwright
import time
import os
import argparse
import re
from pathlib import Path
import sys
import openpyxl
from openpyxl.cell.cell import MergedCell

# =========================================================
# CONFIGURATION
# =========================================================

ROOT_DIR = Path(__file__).resolve().parent

DEFAULT_EXCEL = "Assignment 1 - Test cases.xlsx"

DEFAULT_SHEET_NAME = "Sheet1"

DEFAULT_FRONTEND_URL = "https://www.pixelssuite.com/chat-translator"

DEFAULT_WAIT_MS = 12000
DEFAULT_RETRIES = 15
DEFAULT_RETRY_WAIT_MS = 3000
DEFAULT_TYPE_DELAY_MS = 100
DEFAULT_TIMEOUT_MS = 180000
DEFAULT_SLOW_MO_MS = 300

INPUT_HEADERS = [
    "Input",
    "Singlish",
]

EXPECTED_HEADERS = [
    "Expected Output",
    "Sinhala",
]

ACTUAL_HEADERS = [
    "Actual Output",
]

STATUS_HEADERS = [
    "Status",
]


# =========================================================
# HELPERS
# =========================================================

def configure_stdout():
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass


def normalize(text):
    if text is None:
        return ""

    return re.sub(
        r"[^a-z0-9]",
        "",
        str(text).strip().lower()
    )


def get_headers(ws):

    return [
        ws.cell(row=1, column=i).value
        for i in range(1, ws.max_column + 1)
    ]


def find_column(headers, candidates):

    normalized = {}

    for i, value in enumerate(headers, start=1):

        if value is not None:
            normalized[normalize(value)] = i

    for candidate in candidates:

        key = normalize(candidate)

        if key in normalized:
            return normalized[key]

    return None


def get_merged_cell(ws, row, col):

    cell = ws.cell(row=row, column=col)

    if not isinstance(cell, MergedCell):
        return cell

    for rng in ws.merged_cells.ranges:

        if (
            rng.min_row <= row <= rng.max_row
            and rng.min_col <= col <= rng.max_col
        ):
            return ws.cell(
                row=rng.min_row,
                column=rng.min_col
            )

    return cell


def set_cell(ws, row, col, value):

    cell = get_merged_cell(ws, row, col)
    cell.value = value


def dismiss_popups(page):

    popup_buttons = [
        "Accept",
        "OK",
        "I Agree",
        "Got it",
        "Accept all",
    ]

    for name in popup_buttons:

        try:

            btn = page.get_by_role(
                "button",
                name=name
            ).first

            if btn.is_visible():

                btn.click(timeout=5000)

                page.wait_for_timeout(1000)

        except Exception:
            pass


def clear_textbox(page, locator):

    try:

        locator.click(timeout=5000)

        page.keyboard.press("Control+A")

        page.keyboard.press("Backspace")

        locator.fill("")

    except Exception:
        pass


def read_output(locator):

    methods = [
        lambda: locator.input_value(),
        lambda: locator.inner_text(),
        lambda: locator.text_content(),
    ]

    for method in methods:

        try:

            value = method()

            if value and str(value).strip():
                return str(value).strip()

        except Exception:
            pass

    return ""


# =========================================================
# ARGUMENTS
# =========================================================

def parse_args():

    parser = argparse.ArgumentParser()

    parser.add_argument(
        "--excel",
        default=DEFAULT_EXCEL
    )

    parser.add_argument(
        "--sheet",
        default=DEFAULT_SHEET_NAME
    )

    parser.add_argument(
        "--url",
        default=DEFAULT_FRONTEND_URL
    )

    parser.add_argument(
        "--wait-ms",
        type=int,
        default=DEFAULT_WAIT_MS
    )

    parser.add_argument(
        "--retries",
        type=int,
        default=DEFAULT_RETRIES
    )

    parser.add_argument(
        "--retry-wait-ms",
        type=int,
        default=DEFAULT_RETRY_WAIT_MS
    )

    parser.add_argument(
        "--type-delay-ms",
        type=int,
        default=DEFAULT_TYPE_DELAY_MS
    )

    parser.add_argument(
        "--timeout-ms",
        type=int,
        default=DEFAULT_TIMEOUT_MS
    )

    parser.add_argument(
        "--slow-mo-ms",
        type=int,
        default=DEFAULT_SLOW_MO_MS
    )

    parser.add_argument(
        "--save-every",
        type=int,
        default=1
    )

    parser.add_argument(
        "--keep-open",
        action="store_true"
    )

    return parser.parse_args()


# =========================================================
# MAIN
# =========================================================

def run_test():

    configure_stdout()

    args = parse_args()

    if not os.path.exists(args.excel):

        print("Excel file not found.")
        return

    try:

        wb = openpyxl.load_workbook(args.excel)

    except Exception as e:

        print("Excel loading failed:", e)
        return

    if args.sheet in wb.sheetnames:
        ws = wb[args.sheet]
    else:
        ws = wb.active

    print("Using sheet:", ws.title)

    headers = get_headers(ws)

    print("Headers:", headers)

    input_col = find_column(
        headers,
        INPUT_HEADERS
    )

    expected_col = find_column(
        headers,
        EXPECTED_HEADERS
    )

    actual_col = find_column(
        headers,
        ACTUAL_HEADERS
    )

    status_col = find_column(
        headers,
        STATUS_HEADERS
    )

    if not actual_col:

        actual_col = ws.max_column + 1

        ws.cell(
            row=1,
            column=actual_col
        ).value = "Actual Output"

    if not status_col:

        status_col = ws.max_column + 1

        ws.cell(
            row=1,
            column=status_col
        ).value = "Status"

    print("Input column:", input_col)

    if not input_col:

        print("Input column not found.")
        return

    print(
        f"Starting automation with {ws.max_row - 1} rows..."
    )

    with sync_playwright() as p:

        browser = p.chromium.launch(
            headless=False,
            slow_mo=max(
                300,
                int(args.slow_mo_ms)
            ),
            args=[
                "--disable-blink-features=AutomationControlled",
                "--disable-dev-shm-usage",
                "--no-sandbox",
                "--start-maximized"
            ]
        )

        context = browser.new_context(
            viewport={
                "width": 1366,
                "height": 768
            },
            user_agent=(
                "Mozilla/5.0 "
                "(Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 "
                "(KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            locale="en-US"
        )

        page = context.new_page()

        page.add_init_script("""
        Object.defineProperty(navigator, 'webdriver', {
            get: () => undefined
        });
        """)

        page.set_default_timeout(
            max(
                180000,
                int(args.timeout_ms)
            )
        )

        try:

            print("Opening website...")

            page.goto(
                args.url,
                wait_until="domcontentloaded",
                timeout=args.timeout_ms
            )

            # Wait for page to fully load
            page.wait_for_timeout(5000)

            # Verify we are on the correct page
            current_url = page.url
            print("Current URL:", current_url)

            if "pixelssuite" not in current_url:
                print("Wrong page loaded! Trying again...")
                page.goto(
                    args.url,
                    wait_until="domcontentloaded",
                    timeout=args.timeout_ms
                )
                page.wait_for_timeout(5000)

            page.wait_for_selector(
                "textarea",
                timeout=args.timeout_ms
            )

            print("Frontend loaded successfully.")

        except Exception as e:

            print("Website loading failed:", e)

            browser.close()
            return

        try:

            input_box = page.locator(
                "textarea"
            ).nth(0)

            output_box = page.locator(
                "textarea"
            ).nth(1)

            transliterate_button = page.get_by_role(
                "button",
                name=re.compile(
                    r"Transliterate",
                    re.IGNORECASE
                )
            ).first

            print("UI elements located.")

        except Exception as e:

            print("Locator error:", e)

            browser.close()
            return

        processed = 0

        for row in range(2, ws.max_row + 1):

            print("\n==========================")
            print(f"Checking row {row}")

            try:

                value = get_merged_cell(
                    ws,
                    row,
                    input_col
                ).value

                singlish_input = (
                    str(value).strip()
                    if value else ""
                )

                if not singlish_input:

                    print("Skipped empty row.")
                    continue

                expected_output = ""

                if expected_col:

                    ev = get_merged_cell(
                        ws,
                        row,
                        expected_col
                    ).value

                    expected_output = (
                        str(ev).strip()
                        if ev else ""
                    )

                print("Input:", singlish_input)

                dismiss_popups(page)

                previous_output = read_output(
                    output_box
                )

                clear_textbox(
                    page,
                    input_box
                )

                # Make sure we are on pixelssuite before typing
                if "pixelssuite" not in page.url:
                    print("Page drifted! Reloading...")
                    page.goto(args.url, wait_until="domcontentloaded", timeout=args.timeout_ms)
                    page.wait_for_timeout(5000)
                    page.wait_for_selector("textarea", timeout=args.timeout_ms)

                input_box.click(
                    timeout=10000
                )

                input_box.type(
                    singlish_input,
                    delay=args.type_delay_ms
                )

                page.wait_for_timeout(3000)

                try:

                    transliterate_button.wait_for(
                        timeout=20000
                    )

                    if transliterate_button.is_visible():

                        transliterate_button.click(
                            timeout=20000
                        )

                    else:

                        page.keyboard.press("Enter")

                except Exception:

                    try:
                        page.keyboard.press("Enter")
                    except Exception:
                        pass

                print("Waiting for translation...")

                page.wait_for_timeout(
                    args.wait_ms
                )

                actual_output = ""

                for retry in range(args.retries):

                    current = read_output(
                        output_box
                    )

                    print(
                        f"Retry {retry + 1}: {current}"
                    )

                    if (
                        current
                        and current.strip()
                        and current != previous_output
                    ):

                        actual_output = current
                        break

                    page.wait_for_timeout(
                        args.retry_wait_ms
                    )

                print("Output:", actual_output)

                set_cell(
                    ws,
                    row,
                    actual_col,
                    actual_output
                )

                if expected_output:

                    if actual_output == expected_output:
                        status = "PASS"
                    else:
                        status = "FAIL"

                else:
                    status = "COLLECTED"

                set_cell(
                    ws,
                    row,
                    status_col,
                    status
                )

                print("Status:", status)

                processed += 1

                if (
                    args.save_every > 0
                    and processed % args.save_every == 0
                ):

                    try:
                        wb.save(args.excel)

                    except PermissionError:

                        print(
                            "Close the Excel file before running."
                        )

            except Exception as e:

                print("Row error:", e)

                try:

                    set_cell(
                        ws,
                        row,
                        status_col,
                        "UI Error"
                    )

                    wb.save(args.excel)

                except Exception:
                    pass

        try:

            wb.save(args.excel)

        except Exception as e:

            print("Final save failed:", e)

        print("\nAutomation completed.")

        if args.keep_open:

            print("\nBrowser kept open.")
            print("Press CTRL+C to close.")

            try:

                while True:
                    page.wait_for_timeout(1000)

            except KeyboardInterrupt:

                print("\nClosing browser...")

        browser.close()


# =========================================================
# START
# =========================================================

if __name__ == "__main__":
    run_test()